import json
import logging

import openpyxl as openpyxl
import requests

from enum import Enum
import tqdm


class STATUS(Enum):
    AC = "ACCEPTED"
    CANCEL = "CANCELED"
    PENDING = "PENDING"


# 创建logger对象
logger = logging.getLogger('my_logger')

# 设置日志等级
logger.setLevel(logging.INFO)
# 输出的日志信息格式
formatter = logging.Formatter('%(asctime)s-%(filename)s-line:%(lineno)d-%(levelname)s: %(message)s')

# 写到控制台
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

# 写到文件
console_handler = logging.FileHandler("ac.log", encoding="utf8")
console_handler.setLevel(logging.WARNING)
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)


class People:
    def __init__(self, name, id):
        self.name = name
        self.id = id


def set_status(team_id, status):
    status_str = ""
    if status == STATUS.AC.value:
        status_str = "accept"
    elif status == STATUS.PENDING.value:
        status_str = "waitlist"
    elif status == STATUS.CANCEL.value:
        status_str = "cancel"

    url = f"https://icpc.global/api/team/{team_id}/{status_str}"
    headers = {"authorization": auth, "cookie": cookie}
    r = requests.post(url, headers=headers)
    if r.status_code != 200:
        raise ConnectionError(f"更新队员信息失败，状态码：{r.status_code}")

    pass


class Team:
    def __init__(self, name, id, status, school):
        self.name = name
        self.id = id
        self.members = list()
        self.coaches = list()
        self.status = status
        self.school = school

    def add_member(self, member):
        self.members.append(member)

    def add_coaches(self, member):
        self.coaches.append(member)


def get_members_and_coach(id):
    url = f"https://icpc.global/api/team/members/team/{id}"
    headers = {"authorization": auth, "cookie": cookie}
    r = requests.get(url, headers=headers)
    if r.status_code != 200:
        raise ConnectionError(f"获取队伍成员，状态码：{r.status_code}")
    members = list()
    coaches = list()

    members_origin = r.json()

    for p in members_origin:
        member = People(p['name'], p['personId'])
        if p['role'] == "CONTESTANT":
            members.append(member)
        else:
            coaches.append(member)
    return members, coaches


def get_teams():
    url = f"https://icpc.global/api/team/search/{CONTEST_ID}/all?q=proj:teamId,rank,site,team,country,institution,coachName,status,action%3Bsort:team+asc%3B&page=1&size=1000"
    headers = {"authorization": auth, "cookie": cookie}
    r = requests.get(url, headers=headers)
    if r.status_code != 200:
        raise ConnectionError(f"获取队伍请求失败，状态码：{r.status_code}")
    teams_origin = r.json()

    ac_teams = list()
    cancel_teams = list()
    pending_teams = list()
    for t in tqdm.tqdm(teams_origin):
        team = Team(t['team']['name'], t['id'], t['status'], t['institution']['name'])
        members, coaches = get_members_and_coach(team.id)
        for mem in members:
            team.add_member(mem)
        for couch in coaches:
            team.add_coaches(couch)

        if team.status == STATUS.AC.value:
            ac_teams.append(team)
        elif team.status == STATUS.PENDING.value:
            pending_teams.append(team)
        elif team.status == STATUS.CANCEL.value:
            pending_teams.append(team)
    logger.info(
        f"获取队伍成功，ac队伍数:{len(ac_teams)},pending队伍数{len(pending_teams)},cancel队伍数:{len(cancel_teams)}")

    return ac_teams, pending_teams, cancel_teams


def get_local_teams():
    wb = openpyxl.load_workbook("icpc.xlsx")
    ws = wb.active
    local_teams = list()
    for i in range(1, ws.max_row + 1):
        name = ws.cell(i, TEAM_NAME_COL).value
        member1 = People(ws.cell(i, MEMBER1_NAME_COL).value, 0)
        member2 = People(ws.cell(i, MEMBER2_NAME_COL).value, 0)
        member3 = People(ws.cell(i, MEMBER3_NAME_COL).value, 0)
        coach = People(ws.cell(i, COACH_NAME_COL).value, 0)
        school = ws.cell(i, SCHOOL_NAME_COL).value
        team = Team(name, 0, STATUS.PENDING.value, school)
        team.add_coaches(coach)
        if member1.name is not None and member1.name != "":
            team.add_member(member1)
        if member2.name is not None and member2.name != "":
            team.add_member(member2)
        if member3.name is not None and member3.name != "":
            team.add_member(member3)
        local_teams.append(team)
    return local_teams


def find_team_by_name_and_school(teams, name, school):
    for team in teams:
        if team.name.lower() == name.lower() and team.school.lower() == school.lower():
            return team

    return None


def check_name(name1, name2):
    name1 = name1.lower().split(" ")
    name2 = name2.lower().split(" ")
    if len(name1) != len(name2):
        return False

    for i in name1:
        find = False
        for j in name2:
            if i == j:
                find = True
                break
        if not find:
            return False
    return True

def check_team(local_team, team):
    if len(local_team.members) != len(team.members):
        return False, "ICPC官网队员数量与报名表不一致"
    for mem in team.members:
        find = False
        for m in local_team.members:
            if check_name(m.name, mem.name):
                find = True
                break
        if not find:
            return False, f"ICPC官网队员\"{mem.name}\"无法从报名表找到"

    for coach in local_team.coaches:
        find = False
        for coa in team.coaches:
            if check_name(coa.name,coach.name):
                find = True
                break
        if not find:
            return False, f"报名表中教练\"{coach.name}\"无法从ICPC官网找到"

    return True, ""


def main():
    ac_teams, pending_teams, cancel_teams = get_teams()
    local_teams = get_local_teams()

    for team in tqdm.tqdm(pending_teams):
        if len(team.members) == 0 or len(team.coaches) == 0:
            logger.warning(f"队伍id:{team.id},队名:\"{team.name}\",学校:\"{team.school}\" 队伍未设置队员或教练，取消")
            set_status(team.id, STATUS.CANCEL.value)
            team.status = STATUS.CANCEL.value
            cancel_teams.append(team)
            continue
        if find_team_by_name_and_school(ac_teams, team.name, team.school) is not None:
            logger.warning(f"队伍id:{team.id},队名:\"{team.name}\",学校:\"{team.school}\" 已经存在同名被AC队伍，取消")
            set_status(team.id, STATUS.CANCEL.value)
            team.status = STATUS.CANCEL.value
            cancel_teams.append(team)
            continue

        local_team = find_team_by_name_and_school(local_teams, team.name, team.school)
        if local_team is None:
            logger.warning(
                f"队伍id:{team.id},队名:\"{team.name}\",学校:\"{team.school}\" 无法从报名表查询到对应队伍，已取消")
            set_status(team.id, STATUS.CANCEL.value)
            team.status = STATUS.CANCEL.value
            cancel_teams.append(team)
            continue

        check, message = check_team(local_team, team)
        if check:
            logger.info(
                f"队伍id:{team.id},队名:\"{team.name}\",学校:\"{team.school}\" 信息确认，已AC")
            set_status(team.id, STATUS.AC.value)
            team.status = STATUS.AC.value
            ac_teams.append(team)
        else:
            logger.warning(
                f"队伍id:{team.id},队名:\"{team.name}\",学校:\"{team.school}\" 与报名表信息核对失败:" + message)

    for local_team in local_teams:
        if find_team_by_name_and_school(ac_teams, local_team.name, local_team.school) is None:
            logger.warning(f"报名表队伍，队名:\"{local_team.name}\",学校:\"{local_team.school}\" 目前未在ICPC官网AC")


if __name__ == "__main__":
    CONTEST_ID = 5544  # 比赛id

    # 一下填写对应内容（英文）在excel中的列，从1开始
    TEAM_NAME_COL = 12
    MEMBER1_NAME_COL = 20
    MEMBER2_NAME_COL = 29
    MEMBER3_NAME_COL = 38
    COACH_NAME_COL = 16
    SCHOOL_NAME_COL = 3
    with open("authorization.txt") as f:
        auth = f.read()
    with open("cookie.txt") as f:
        cookie = f.read()

    main()
